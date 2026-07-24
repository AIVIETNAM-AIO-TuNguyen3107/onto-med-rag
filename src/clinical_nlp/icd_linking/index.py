from __future__ import annotations

import hashlib
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook

from clinical_nlp.normalization import normalize_search
from clinical_nlp.schemas import LinkCandidate


CODE_RE = re.compile(r"^[A-Z][0-9]{2}(?:\.[0-9A-Z]{1,4})?$")
INDEX_SCHEMA_VERSION = 2
CURATED_ALIASES = {
    "thiếu men g6pd": {"D55.0"},
    "thiếu máu": {"D64.9"},
}


@dataclass(frozen=True)
class ICDConcept:
    code: str
    names: tuple[str, ...]
    models: tuple[str, ...] = ()
    hierarchies: tuple[str, ...] = ()
    is_leaf: bool | None = None


class ICDIndex:
    def __init__(
        self,
        concepts: dict[str, ICDConcept],
        source_sha256: str | None = None,
    ) -> None:
        self.concepts = concepts
        self.source_sha256 = source_sha256
        aliases: dict[str, set[str]] = defaultdict(set)
        for code, concept in concepts.items():
            for name in concept.names:
                key = normalize_search(name)
                if key:
                    aliases[key].add(code)
        for alias, codes in CURATED_ALIASES.items():
            aliases[alias].update(code for code in codes if code in concepts)
        self.alias_to_codes = dict(aliases)
        token_aliases: dict[str, set[str]] = defaultdict(set)
        for alias in self.alias_to_codes:
            for token in set(alias.split()):
                token_aliases[token].add(alias)
        self.token_aliases = dict(token_aliases)

    @classmethod
    def from_workbook(cls, path: Path) -> "ICDIndex":
        digest = _sha256(path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            names_by_code: dict[str, set[str]] = defaultdict(set)
            for row in sheet.iter_rows(min_row=6, values_only=True):
                raw_code = "" if row[1] is None else str(row[1]).strip()
                name = "" if row[2] is None else str(row[2]).strip()
                code = raw_code.rstrip("*†")
                if not CODE_RE.fullmatch(code) or not name:
                    continue
                names_by_code[code].add(name)
        finally:
            workbook.close()
        concepts = {
            code: ICDConcept(code=code, names=tuple(sorted(names)))
            for code, names in names_by_code.items()
        }
        return cls(concepts=concepts, source_sha256=digest)

    @classmethod
    def from_catalog(cls, path: Path) -> "ICDIndex":
        digest = _sha256(path)
        rows: list[dict] = []
        path_to_name: dict[str, str] = {}
        with path.open(encoding="utf-8") as handle:
            for line_number, line in enumerate(handle, start=1):
                if not line.strip():
                    continue
                try:
                    row = json.loads(line)
                except json.JSONDecodeError as exc:
                    raise ValueError(
                        f"invalid ICD catalog JSON at line {line_number}"
                    ) from exc
                if not isinstance(row, dict):
                    raise ValueError(
                        f"ICD catalog row at line {line_number} must be an object"
                    )
                rows.append(row)
                path_value = row.get("path")
                name_value = row.get("name_vi")
                if isinstance(path_value, str) and isinstance(name_value, str):
                    if path_value.strip() and name_value.strip():
                        path_to_name[path_value.strip()] = name_value.strip()

        names_by_code: dict[str, set[str]] = defaultdict(set)
        models_by_code: dict[str, set[str]] = defaultdict(set)
        hierarchies_by_code: dict[str, set[str]] = defaultdict(set)
        leaf_by_code: dict[str, list[bool]] = defaultdict(list)
        for row in rows:
            model = row.get("model")
            if model not in {"type", "disease"}:
                continue
            raw_code = row.get("code")
            raw_name = row.get("name_vi")
            if not isinstance(raw_code, str) or not isinstance(raw_name, str):
                raise ValueError("ICD catalog row has invalid code/name")
            code = raw_code.strip().rstrip("*†")
            name = raw_name.strip()
            if not CODE_RE.fullmatch(code) or not name:
                continue
            names_by_code[code].add(name)
            models_by_code[code].add(model)
            hierarchy = _hierarchy_for(row.get("path"), path_to_name)
            if hierarchy:
                hierarchies_by_code[code].add(hierarchy)
            if isinstance(row.get("is_leaf"), bool):
                leaf_by_code[code].append(row["is_leaf"])
        concepts = {
            code: ICDConcept(
                code=code,
                names=tuple(sorted(names)),
                models=tuple(sorted(models_by_code[code])),
                hierarchies=tuple(sorted(hierarchies_by_code[code])),
                is_leaf=(
                    all(leaf_by_code[code])
                    if leaf_by_code.get(code)
                    else None
                ),
            )
            for code, names in names_by_code.items()
        }
        return cls(concepts=concepts, source_sha256=digest)

    @classmethod
    def from_source(cls, path: Path) -> "ICDIndex":
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            return cls.from_workbook(path)
        if suffix == ".jsonl":
            return cls.from_catalog(path)
        raise ValueError(
            f"unsupported ICD source format {suffix!r}; expected .xlsx, .xlsm, or .jsonl"
        )

    @classmethod
    def load(cls, path: Path) -> "ICDIndex":
        raw = json.loads(path.read_text("utf-8"))
        if raw.get("schema_version") != INDEX_SCHEMA_VERSION:
            raise ValueError("ICD index schema is stale and must be rebuilt")
        concepts = {
            row["code"]: ICDConcept(
                code=row["code"],
                names=tuple(row["names"]),
                models=tuple(row.get("models", [])),
                hierarchies=tuple(row.get("hierarchies", [])),
                is_leaf=row.get("is_leaf"),
            )
            for row in raw["concepts"]
        }
        return cls(concepts=concepts, source_sha256=raw.get("source_sha256"))

    def save(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "schema_version": INDEX_SCHEMA_VERSION,
            "source_sha256": self.source_sha256,
            "concepts": [
                {
                    "code": concept.code,
                    "names": list(concept.names),
                    "models": list(concept.models),
                    "hierarchies": list(concept.hierarchies),
                    "is_leaf": concept.is_leaf,
                }
                for concept in sorted(self.concepts.values(), key=lambda row: row.code)
            ],
        }
        path.write_text(
            json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )

    def contains(self, code: str) -> bool:
        return code in self.concepts

    def retrieve(self, mention: str, limit: int = 10) -> list[LinkCandidate]:
        query = normalize_search(mention)
        if not query:
            return []
        exact = self.alias_to_codes.get(query, set())
        results: dict[str, LinkCandidate] = {}
        for code in exact:
            concept = self.concepts[code]
            score = 0.88 if code.startswith("Z") else 1.0
            results[code] = LinkCandidate(
                identifier=code,
                name=concept.names[0],
                terminology_type=_terminology_type(concept),
                score=score,
                retrieval_sources=["exact"],
                component_scores={
                    "lexical": 1.0,
                    "context_penalty": -0.12 if code.startswith("Z") else 0.0,
                },
                metadata=_metadata(concept),
            )
        if len(results) >= limit:
            return sorted(results.values(), key=lambda row: (-row.score, row.identifier))[
                :limit
            ]

        query_tokens = set(query.split())
        candidate_aliases: set[str] = set()
        for token in query_tokens:
            candidate_aliases.update(self.token_aliases.get(token, ()))
        scored_aliases: list[tuple[float, str, set[str]]] = []
        for alias in candidate_aliases:
            codes = self.alias_to_codes[alias]
            alias_tokens = set(alias.split())
            token_score = (
                len(query_tokens & alias_tokens) / len(query_tokens | alias_tokens)
                if query_tokens | alias_tokens
                else 0.0
            )
            if token_score < 0.25 and query not in alias and alias not in query:
                continue
            char_score = SequenceMatcher(None, query, alias).ratio()
            score = 0.55 * char_score + 0.45 * token_score
            if score >= 0.48:
                scored_aliases.append((score, alias, codes))
        scored_aliases.sort(key=lambda row: (-row[0], len(row[1])))
        for score, alias, codes in scored_aliases:
            for code in codes:
                adjusted_score = max(0.0, score - (0.12 if code.startswith("Z") else 0))
                existing = results.get(code)
                if existing is None or adjusted_score > existing.score:
                    concept = self.concepts[code]
                    results[code] = LinkCandidate(
                        identifier=code,
                        name=concept.names[0],
                        terminology_type=_terminology_type(concept),
                        score=adjusted_score,
                        retrieval_sources=["fuzzy"],
                        component_scores={
                            "lexical": score,
                            "context_penalty": (
                                -0.12 if code.startswith("Z") else 0.0
                            ),
                        },
                        metadata=_metadata(concept),
                    )
            if len(results) >= limit * 3:
                break
        return sorted(results.values(), key=lambda row: (-row.score, row.identifier))[
            :limit
        ]

    def all_aliases(self, min_length: int = 4) -> list[tuple[str, str]]:
        rows: list[tuple[str, str]] = []
        for concept in self.concepts.values():
            for name in concept.names:
                if len(name) >= min_length:
                    rows.append((name, concept.code))
        return rows


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _hierarchy_for(raw_path: object, path_to_name: dict[str, str]) -> str:
    if not isinstance(raw_path, str) or not raw_path:
        return ""
    prefixes: list[str] = []
    parts = raw_path.split("/")
    for index in range(1, len(parts) + 1):
        prefix = "/".join(parts[:index])
        name = path_to_name.get(prefix)
        if name and (not prefixes or prefixes[-1] != name):
            prefixes.append(name)
    return " > ".join(prefixes)


def _terminology_type(concept: ICDConcept) -> str:
    suffix = ",".join(concept.models)
    return f"ICD10:{suffix}" if suffix else "ICD10"


def _metadata(concept: ICDConcept) -> dict[str, object]:
    return {
        "models": list(concept.models),
        "hierarchies": list(concept.hierarchies),
        "is_leaf": concept.is_leaf,
    }
